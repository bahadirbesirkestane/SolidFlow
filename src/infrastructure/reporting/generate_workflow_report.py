import json
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def main():
    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Ozet"
    rows_sheet = workbook.create_sheet("Is Akisi")

    write_summary(summary_sheet, payload)
    write_rows(rows_sheet, payload.get("rows", []))

    workbook.save(output_path)


def write_summary(sheet, payload):
    summary = payload.get("summary", {})
    header_fill = PatternFill(fill_type="solid", fgColor="0B6E61")
    header_font = Font(color="FFFFFF", bold=True)

    sheet["A1"] = "Alan"
    sheet["B1"] = "Deger"
    for cell in sheet[1]:
      cell.fill = header_fill
      cell.font = header_font

    summary_rows = [
        ("Taranan Klasor", payload.get("scannedFolder", "")),
        ("Toplam Dosya", summary.get("totalFiles", 0)),
        ("Surec Atanmis", summary.get("assignedFiles", 0)),
        ("Belirsiz", summary.get("uncertainFiles", 0)),
    ]

    row_index = 2
    for label, value in summary_rows:
        sheet[f"A{row_index}"] = label
        sheet[f"B{row_index}"] = value
        row_index += 1

    row_index += 1
    sheet[f"A{row_index}"] = "Surec Dagilimi"
    sheet[f"A{row_index}"].font = Font(bold=True)
    row_index += 1

    for process, count in OrderedDict(sorted(summary.get("byProcess", {}).items())).items():
        sheet[f"A{row_index}"] = process
        sheet[f"B{row_index}"] = count
        row_index += 1

    row_index += 1
    sheet[f"A{row_index}"] = "Hizmet Dagilimi"
    sheet[f"A{row_index}"].font = Font(bold=True)
    row_index += 1

    for service_type, count in OrderedDict(sorted(summary.get("byServiceType", {}).items())).items():
        sheet[f"A{row_index}"] = service_type
        sheet[f"B{row_index}"] = count
        row_index += 1

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 42


def write_rows(sheet, rows):
    columns = [
        "Parca Kodu",
        "Dosya Adi",
        "Dosya Tipi",
        "Ana Grup",
        "Surec",
        "Hizmet",
        "Guven",
        "Eslesme",
        "Klasor",
        "Goreli Yol",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="1F2933")
    header_font = Font(color="FFFFFF", bold=True)

    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = column
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row in enumerate(rows, start=2):
        values = [
            row.get("partCode", ""),
            row.get("fileName", ""),
            row.get("fileType", ""),
            row.get("mainGroup", ""),
            row.get("suggestedProcess", ""),
            row.get("serviceType", ""),
            row.get("confidence", ""),
            row.get("matchedBy", ""),
            row.get("folder", ""),
            row.get("relativePath", ""),
        ]

        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index).value = value

    widths = [14, 34, 18, 18, 18, 22, 14, 24, 30, 44]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


if __name__ == "__main__":
    main()
