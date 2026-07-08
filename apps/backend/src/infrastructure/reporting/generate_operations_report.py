import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def main():
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    output_format = sys.argv[3]

    with open(input_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    if output_format == "xlsx":
        write_workbook(payload, output_path)
        return

    if output_format == "pdf":
        write_pdf(payload, output_path)
        return

    raise ValueError(f"Unsupported format: {output_format}")


def write_workbook(payload, output_path):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Ozet"
    workflows_sheet = workbook.create_sheet("Workflowlar")
    steps_sheet = workbook.create_sheet("Adimlar")
    open_jobs_sheet = workbook.create_sheet("Acik Isler")
    audit_sheet = workbook.create_sheet("Audit")

    write_summary_sheet(summary_sheet, payload)
    write_workflows_sheet(workflows_sheet, payload.get("workflows", []))
    write_steps_sheet(steps_sheet, payload.get("workflows", []))
    write_open_jobs_sheet(open_jobs_sheet, payload.get("openJobs", []))
    write_audit_sheet(audit_sheet, payload.get("auditEvents", []))

    workbook.save(output_path)


def write_summary_sheet(sheet, payload):
    project = payload.get("project", {})
    progress = payload.get("progress", {})
    rows = [
        ("Proje Kodu", project.get("code", "")),
        ("Proje Adi", project.get("name", "")),
        ("Aciklama", project.get("description", "")),
        ("Toplam Workflow", progress.get("totalInstances", 0)),
        ("Toplam Adim", progress.get("totalSteps", 0)),
        ("Tamamlanan Adim", progress.get("completedSteps", 0)),
        ("Ilerleme", f"%{progress.get('completionPercentage', 0)}"),
        ("Acik Is", len(payload.get("openJobs", []))),
        ("Audit Kaydi", len(payload.get("auditEvents", []))),
        ("Rapor Tarihi", payload.get("generatedAt", "")),
    ]
    write_key_value_sheet(sheet, rows)


def write_workflows_sheet(sheet, workflows):
    rows = [[
        "Workflow",
        "Template",
        "Kalem",
        "Adim Sayisi",
        "Durum",
        "Ilerleme",
        "Aktif Adim",
    ]]
    for workflow in workflows:
        current_step = workflow.get("currentStep") or {}
        rows.append([
            workflow.get("name", ""),
            workflow.get("templateName", ""),
            workflow.get("itemLabel", ""),
            len(workflow.get("steps", [])),
            workflow.get("status", ""),
            f"%{workflow.get('progressPercent', 0)}",
            current_step.get("name", ""),
        ])
    write_table_sheet(sheet, rows, [32, 24, 22, 14, 16, 14, 28])


def write_steps_sheet(sheet, workflows):
    rows = [[
        "Workflow",
        "Sira",
        "Adim",
        "Durum",
        "Atananlar",
        "Opsiyonel",
        "Onaylayan",
        "Devir",
        "Not",
        "Tamamlanma",
    ]]
    for workflow in workflows:
        for step in workflow.get("steps", []):
            rows.append([
                workflow.get("name", ""),
                step.get("sequenceNo", ""),
                step.get("name", ""),
                step.get("status", ""),
                ", ".join(step.get("assigneeIds", []) or []) or step.get("assignee", ""),
                "Evet" if step.get("isOptional") else "Hayir",
                step.get("approvedBy", ""),
                step.get("handoverTo", ""),
                step.get("completionNote", ""),
                step.get("completedAt", ""),
            ])
    write_table_sheet(sheet, rows, [30, 8, 22, 14, 28, 10, 18, 18, 34, 20])


def write_open_jobs_sheet(sheet, jobs):
    rows = [["Baslik", "Durum", "Kaynak", "Aciklama", "Olusturulma"]]
    for job in jobs:
        rows.append([
            job.get("title", ""),
            job.get("status", ""),
            job.get("sourceType", ""),
            job.get("description", ""),
            job.get("createdAt", ""),
        ])
    write_table_sheet(sheet, rows, [34, 14, 18, 50, 20])


def write_audit_sheet(sheet, audit_events):
    rows = [["Tarih", "Islem", "Varlik Tipi", "Varlik Id", "Ozet"]]
    for event in audit_events:
        rows.append([
            event.get("createdAt", ""),
            event.get("action", ""),
            event.get("entityType", ""),
            event.get("entityId", ""),
            summarize_payload(event.get("payload", {})),
        ])
    write_table_sheet(sheet, rows, [20, 18, 18, 26, 60])


def write_key_value_sheet(sheet, rows):
    header_fill = PatternFill(fill_type="solid", fgColor="0B6E61")
    header_font = Font(color="FFFFFF", bold=True)
    sheet["A1"] = "Alan"
    sheet["B1"] = "Deger"
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, (label, value) in enumerate(rows, start=2):
        sheet[f"A{index}"] = label
        sheet[f"B{index}"] = value
        sheet[f"A{index}"].alignment = Alignment(vertical="top", wrap_text=True)
        sheet[f"B{index}"].alignment = Alignment(vertical="top", wrap_text=True)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 48


def write_table_sheet(sheet, rows, widths):
    header_fill = PatternFill(fill_type="solid", fgColor="1F2933")
    header_font = Font(color="FFFFFF", bold=True)

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == 1:
                cell.fill = header_fill
                cell.font = header_font

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def write_pdf(payload, output_path):
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    section_style = styles["Heading2"]
    small_style = styles["BodyText"]
    small_style.fontName = "Helvetica"
    small_style.fontSize = 9
    muted_style = ParagraphStyle(
        "Muted",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#5b6470"),
        leading=11,
    )

    document = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    story = []
    project = payload.get("project", {})
    progress = payload.get("progress", {})

    story.append(Paragraph(f"Operasyonel Workflow Raporu - {project.get('code', '')}", title_style))
    story.append(Paragraph(project.get("name", ""), section_style))
    story.append(Paragraph(project.get("description", "") or "Aciklama girilmemis.", muted_style))
    story.append(Spacer(1, 6))

    summary_table = Table([
        ["Toplam Workflow", "Toplam Adim", "Tamamlanan Adim", "Ilerleme", "Acik Is", "Audit Kaydi"],
        [
            str(progress.get("totalInstances", 0)),
            str(progress.get("totalSteps", 0)),
            str(progress.get("completedSteps", 0)),
            f"%{progress.get('completionPercentage', 0)}",
            str(len(payload.get("openJobs", []))),
            str(len(payload.get("auditEvents", []))),
        ],
    ], colWidths=[34 * mm, 28 * mm, 32 * mm, 22 * mm, 22 * mm, 24 * mm])
    summary_table.setStyle(base_table_style())
    story.append(summary_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Workflow Agaci", section_style))
    story.append(Paragraph(
        "Bu alandaki her blok bir is akisidir. Siralar yukaridan asagi okunur; ok yapisi mantiksal is akisini gosterir.",
        muted_style,
    ))
    story.append(Spacer(1, 6))

    for workflow in payload.get("workflows", []):
        story.append(Paragraph(f"{workflow.get('name', '')} - %{workflow.get('progressPercent', 0)}", styles["Heading3"]))
        step_lines = []
        for step in workflow.get("steps", []):
            assignees = ", ".join(step.get("assigneeIds", []) or []) or step.get("assignee", "") or "Atama yok"
            step_lines.append(
                f"{step.get('sequenceNo', '')}. {step.get('name', '')}  ->  {step.get('status', '')}  |  Atanan: {assignees}"
            )
        flow_text = "<br/>".join(step_lines) if step_lines else "Bu workflow icinde adim yok."
        story.append(Paragraph(flow_text, small_style))
        story.append(Spacer(1, 8))

    if payload.get("openJobs"):
        story.append(Paragraph("Acik Isler", section_style))
        open_job_rows = [["Baslik", "Durum", "Aciklama"]]
        for job in payload.get("openJobs", [])[:12]:
            open_job_rows.append([
                job.get("title", ""),
                job.get("status", ""),
                job.get("description", ""),
            ])
        open_job_table = Table(open_job_rows, colWidths=[80 * mm, 22 * mm, 150 * mm])
        open_job_table.setStyle(base_table_style())
        story.append(open_job_table)
        story.append(Spacer(1, 8))

    if payload.get("auditEvents"):
        story.append(Paragraph("Audit Akisi Ozet", section_style))
        audit_rows = [["Tarih", "Islem", "Ozet"]]
        for event in payload.get("auditEvents", [])[:12]:
            audit_rows.append([
                event.get("createdAt", ""),
                event.get("action", ""),
                summarize_payload(event.get("payload", {})),
            ])
        audit_table = Table(audit_rows, colWidths=[42 * mm, 30 * mm, 180 * mm])
        audit_table.setStyle(base_table_style())
        story.append(audit_table)

    document.build(story)


def base_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B6E61")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d8cfbf")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f8f4ed")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


def summarize_payload(payload):
    if not isinstance(payload, dict):
        return ""

    parts = []
    for key, value in list(payload.items())[:5]:
        if isinstance(value, list):
            normalized_value = ", ".join(str(item) for item in value[:4])
        else:
            normalized_value = str(value)
        parts.append(f"{key}: {normalized_value}")
    return " | ".join(parts)


if __name__ == "__main__":
    main()
